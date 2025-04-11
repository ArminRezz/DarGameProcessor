import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

"""
Sports Game Processor
--------------------
This script processes sports game data from CSV files and generates formatted Excel reports.
It categorizes games as 'Vegas' or 'Public' based on various criteria including sport type,
day of the week, and game timing.

Main features:
- Processes CSV input files containing game information
- Categorizes games based on sport-specific rules
- Generates formatted Excel output with conditional formatting
- Supports multiple sports (NBA, NHL, MLB, NCAAB, NCAAF, NFL)
"""

# Helper Functions
# ---------------

def title_case(text):
    """Converts text to title case, handling None/empty values."""
    return text.strip().title() if text else ""

def clean_time(time_str):
    """Standardizes time string format (e.g., '7:30 PM EST' -> '7:30 PM')."""
    if not time_str:
        return ""
    # Remove all timezone indicators and extra spaces
    time_str = time_str.strip().lower()
    for tz in ['est', 'edt', 'cst', 'cdt', 'mst', 'mdt', 'pst', 'pdt']:
        time_str = time_str.replace(tz, '')
    # Ensure proper spacing for AM/PM
    time_str = time_str.replace('pm', ' PM').replace('am', ' AM')
    return time_str.strip().upper()

def convert_to_24h(time_str):
    """Converts 12-hour time format to 24-hour decimal format (e.g., '7:30 PM' -> 19.5)."""
    if not time_str:
        return None
    
    # Clean up the time string
    time_str = time_str.strip().lower()
    
    # Remove any remaining letters except 'am'/'pm'
    cleaned = ''.join(c for c in time_str if c.isdigit() or c in ':amp ')
    
    try:
        # Split time and meridiem
        time_part = cleaned.replace('am', '').replace('pm', '').strip()
        is_pm = 'pm' in time_str.lower()
        
        if ':' in time_part:
            hour, minute = map(int, time_part.split(':'))
        else:
            hour = int(time_part)
            minute = 0
            
        # Convert to 24-hour format
        if is_pm and hour < 12:
            hour += 12
        elif not is_pm and hour == 12:
            hour = 0
            
        return hour + (minute / 60)
    except ValueError:
        print(f"Warning: Could not parse time string: {time_str}")
        return None

# Game Classification Functions
# ---------------------------

def get_day_type(sport, day):
    """Determines if a given day is typically a 'vegas' or 'public' day for a sport."""
    day = day.lower()
    vegas_days = {
        'nba': ['tuesday', 'thursday', 'saturday', 'sunday'],
        'nhl': ['tuesday', 'thursday', 'saturday', 'sunday'],
        'mlb': ['tuesday', 'thursday', 'saturday', 'sunday'],
        'ncaab': ['tuesday', 'thursday', 'saturday', 'sunday'],
        'ncaaf': ['thursday', 'friday', 'saturday afternoon'],
    }
    if sport in vegas_days:
        return 'vegas' if any(day.startswith(d) for d in vegas_days[sport]) else 'public'
    return 'unknown'

def determine_game_type(sport, date, game_index, start_time=None, time_to_game_type=None):
    """
    Determines whether a game is 'Vegas' or 'Public' based on multiple factors:
    - Sport type
    - Day of the week
    - Start time
    - Game sequence
    """
    sport = sport.lower()
    day = date.split('-')[2]
    day_type = get_day_type(sport, day)

    if sport in ['nba', 'nhl', 'ncaab']:
        # Ensure games at the same time share the same type
        if start_time in time_to_game_type:
            return time_to_game_type[start_time]

        game_type = 'Vegas' if (day_type == 'public' and game_index % 2 == 0) or (day_type == 'vegas' and game_index % 2 == 1) else 'Public'
        time_to_game_type[start_time] = game_type
        return game_type

    if sport == 'nfl':
        if '1 pm' in date:
            return 'Public'
        elif '4 pm' in date or 'sunday night' in date or 'thursday' in date:
            return 'Vegas'
        elif 'monday' in date:
            return 'Public'
        else:
            return 'Unknown'

    if sport == 'mlb' and 'wednesday' in date and start_time is not None:
        return 'Public' if start_time < 17.4 else 'Vegas'

    if sport == 'ncaaf':
        if 'saturday early' in date:
            return 'Public'
        elif 'saturday afternoon' in date:
            return 'Vegas'
        elif 'saturday night' in date:
            return 'Mixed'
        return 'Vegas'

    return day_type.title()

# Main Processing Function
# -----------------------

def process_to_excel(input_file):
    """
    Main function to process sports game data from CSV to formatted Excel.
    
    Args:
        input_file (str): Path to input CSV file (format: 'sport_date.csv')
    
    Returns:
        None: Saves formatted Excel file to disk
    """
    # Extract parts of the filename: sport and date
    filename = input_file.split('.')[0]
    sport, date = filename.split('_')

    # Read the CSV for that day and sport
    df = pd.read_csv(input_file)

    processed_rows = []
    time_to_game_type = {} # Dictionary to track games at the same time
    for i, row in df.iterrows():
        start_time_str = clean_time(row['start_time'])
        start_time_float = convert_to_24h(start_time_str)

        # Determine the game type based on the sport, date, and game index
        game_type = determine_game_type(sport, date, i, start_time_float, time_to_game_type)

        processed_rows.append({
            'Sport': sport.upper(),
            'Date': date,
            'Start Time': start_time_str,
            'Matchup': row['teams'].title(),
            'Game Type': game_type
        })

    # Create output DataFrame and save to Excel
    output_df = pd.DataFrame(processed_rows)
    output_file = f"{sport}_{date}.xlsx"
    output_df.to_excel(output_file, index=False, sheet_name='Games', engine='openpyxl')

    # Excel styling and conditional formatting
    wb = load_workbook(output_file)
    ws = wb['Games']

    # Apply conditional formatting (red for Vegas, blue for Public)
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid") # Vegas = Red
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Public = Blue

    # Apply to Game Type column
    ws.conditional_formatting.add(f"E2:E{ws.max_row}", FormulaRule(formula=[f'E2="Vegas"'], fill=red_fill))
    ws.conditional_formatting.add(f"E2:E{ws.max_row}", FormulaRule(formula=[f'E2="Public"'], fill=blue_fill))

    wb.save(output_file)

    print(f"Processing complete! The file has been saved as: {output_file}")

# Script Execution
# ---------------

if __name__ == "__main__":
    # Example usage
    process_to_excel('nba_2025-04-11.csv')